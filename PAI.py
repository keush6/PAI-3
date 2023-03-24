# -*- coding: utf-8 -*-
"""
Created on Tue Nov 29 14:56:47 2022

"""

### PAI N°3 : Programme python ###

### Importation des modules ###

import imaplib, email,datetime,mailbox
from tkinter import * 
import tkinter as tk
import sqlite3
import openpyxl as xl
import time
from PIL import ImageTk,Image
from datetime import datetime
from pytz import timezone

### Récupération du corps des mails ###


def connexion(servername): 
    #gestion des mot de passe et user (introduire une table de hashage (voir double table pour plus de sécurité))
    ORG_EMAIL = "@outlook.fr" 
    usernm = "test.pai3" + ORG_EMAIL 
    passwd = "Tomblanchard3."
    conn = imaplib.IMAP4_SSL(servername)
    conn.login(usernm,passwd)
    conn.select('Inbox')
    result, data = conn.uid('search', None, "UNSEEN") # (ALL/UNSEEN)
    i = len(data[0].split())
    return(i,data,conn)


### Interface Graphique (choix des paramètres) ###        
class FenPrincipale(Tk):
    ### Action à rélaiser ne fonction du type de mail ###
    def plan_de_vol(self,corps,id_aeronef):                 # Fonction terminée fonctionelle 
        conn = sqlite3.connect(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols_pai_3.db')
        cur = conn.cursor()

        # Identifiant aéronef
        cur.execute('''REPLACE INTO "Plans de vols"(Aeronef) VALUES (?)''',(id_aeronef,))
        
        # Identifiant aérodrome de départ
        ligne=corps[4].split('-')
        depart=ligne[1]
        cur.execute('''UPDATE "Plans de vols" SET "Aerodrome de depart" = ? WHERE Aeronef = ?''',[(depart[0:5]),id_aeronef])

        # Heure de départ
        A=depart[5:10]
        B = A[0:3] + ':' + A[3:5]

        cur.execute('''UPDATE "Plans de vols" SET "Heure de départ" = ? WHERE Aeronef = ?''',[(B),id_aeronef])

        # Identifiant aérodrome d'arrivée
        ligne2=corps[8].split('-')
        arrivee=ligne2[1]
        cur.execute('''UPDATE "Plans de vols" SET "Aerodrome d'arrivee" = ? WHERE Aeronef = ?''',[(arrivee[0:5]),id_aeronef])

        # Durée du vol
        C=arrivee[5:10]
        D = C[0:3] + ':' + C[3:5]
        cur.execute('''UPDATE "Plans de vols" SET "Duree du vol" = ? WHERE Aeronef = ?''',[(D),id_aeronef])

        # Heure d'arrivée
        heure=int(depart[6:8])+int(arrivee[6:8])
        minute=int(depart[8:10])+int(arrivee[8:10])

        if int(minute)>60:
            minute=int(minute)-60
            heure+=1
        heure_arrivee = str(heure)+str(minute)

        E = heure_arrivee[0:2] + ':' + heure_arrivee[2:4]
        
        cur.execute('''UPDATE "Plans de vols" SET "Heure d'arrivee" = ? WHERE Aeronef = ?''',[(E),id_aeronef])

        # Chemin
        
        ligne3 = corps[6].split(' ')
        ligne4 = ligne3[2:len(ligne3)]
        villes = ' '.join(ligne4)

        cur.execute('''UPDATE "Plans de vols" SET "Chemin" = ? WHERE Aeronef = ?''',[(villes),id_aeronef])

        print("déclaration de plan de vol")
        conn.commit()
        conn.close()

    def ecriture_excel(self,corps, id_aeronef):             # Fonction terminée fonctionelle 
        ### Fonction qui inscrit le mail dans le fichier Excel ###

        #Ouverture du fichier
        wb = xl.load_workbook(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
        feuille = wb['Vols en cours']

        #Ligne excel
        i=6
        while feuille.cell(i, 4).value != None :
            i+=1

        # Identifiant aérodrome de départ
        ligne=corps[4].split('-')
        depart=ligne[1]

        #Recuperation vol dans bdd
        conn = sqlite3.connect(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols_pai_3.db')
        cur = conn.cursor()
        cur.execute('''SELECT "Heure de départ","Duree du vol", "Aerodrome d'arrivee", "Heure d'arrivee", "Chemin" FROM "Plans de vols" WHERE Aeronef = ? ''', (id_aeronef,))
        
        vol=[]
        for ligne in cur.fetchall():
            vol=list(ligne)
        

        cur.close()
        conn.close()

        #Ecriture dans le fichier excel
        feuille.cell(i,4).value = id_aeronef
        feuille.cell(i,5).value = depart[0:5]
        feuille.cell(i,6).value = vol[0]
        feuille.cell(i,7).value = vol[1]
        feuille.cell(i,8).value = vol[2]
        feuille.cell(i,9).value = vol[3]
        feuille.cell(i,10).value = vol[4]

        #Sauvegarder
        wb.save(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
               
    def message_delai(self,corps,id_aeronef):               # Fonction terminée à tester

        #Base de donnée
        ligne=corps[4].split('-')
        depart=ligne[1]
        
        conn = sqlite3.connect(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols_pai_3.db')
        cur = conn.cursor()
        cur.execute('''UPDATE "Plans de vols" SET "Heure de départ" = ? WHERE Aeronef = ? AND "Aerodrome de depart" = ?''', (depart[5:10],id_aeronef,depart[0:5]))
        
        ligne2=corps[8].split('-')
        arrivee=ligne2[1]

        heure=int(depart[6:8])+int(arrivee[6:8])
        minute=int(depart[8:10])+int(arrivee[8:10])

        if int(minute)>60:
            minute=int(minute)-60
            heure+=1
        heure_arrivee = str(heure)+str(minute)
        
        cur.execute('''UPDATE "Plans de vols" SET "Heure d'arrivee" = ? WHERE Aeronef = ? AND "Aerodrome de depart" = ?''', (depart[5:10],id_aeronef,depart[0:5]))

        conn.commit()
        conn.close()

        #Excel   
        wb = xl.load_workbook(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
        feuille = wb['Vols en cours']

        
        for row in feuille.iter_rows():
             for cell in row:
                 if cell.value == id_aeronef:
                     a = (cell.row,cell.column)
        
        feuille.cell(row=a[0],column=6).value=depart[5:10]

        feuille.cell(row=a[0],column=9).value=heure_arrivee

        wb.save(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
        
    def message_changement(self,corps,id_aeronef):
        ligne=corps[0].split('-')
        depart = ligne[2]
        arrivee = ligne[3]
        
        conn = sqlite3.connect('/Users/thibautdejean/Desktop/vols_pai.db')
        cur = conn.cursor()
        cur.execute("INSERT INTO Plans_de_Vols(Heure_de_départ) VALUE (?) WHERE Aeronef = (?)", (depart[4:8],id_aeronef))
        cur.execute("INSERT INTO Plans_de_Vols(Aerodrome_de_départ) VALUE (?) WHERE Aeronef = (?)", (depart[0:4],id_aeronef))
        cur.execute("INSERT INTO Plans_de_Vols(Heure_d'arrivee) VALUE (?) WHERE Aeronef = (?)", (arrivee,id_aeronef))
        
        conn.close()
 
    def message_annulation(self,corps,id_aeronef):          # Fonction terminée à tester
        #Base de donnée
        conn = sqlite3.connect(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols_pai_3.db')
        cur = conn.cursor()

        cur.execute('''DELETE FROM "Plans de vols" WHERE Aeronef = ?''', (id_aeronef,))

        conn.commit()
        conn.close()

        #Fichier excel
        wb = xl.load_workbook(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
        feuille = wb['Vols en cours']

        
        for row in feuille.iter_rows():
             for cell in row:
                 if cell.value == id_aeronef:
                     a = (cell.row,cell.column)
        
        for j in range(4,11):
            feuille.cell(row = a[0], column = j).value = None
            fill = xl.PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
            feuille.cell(row = a[0], column = j).fill = fill

        wb.save(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
                       
    def message_depart(self,corps,id_aeronef): # Fonction terminée fonctionnelle 
               
        # Identification de l'aeronef
        ligne=corps[0].split('-')
        b = ligne[1].split(' ')
        id = ' '+b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]+' '
        print(id)

    
        # CHangement de couleur sur l'excel
        wb = xl.load_workbook(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
        feuille = wb['Vols en cours']

        
        
        a=[1,1]
        for row in feuille.iter_rows():
             for cell in row :
                 if str(cell.value) == str(id) : 
                    a = (cell.column,cell.row)
                    print(a)

         
        for j in range(4,11):
            fill = xl.styles.PatternFill(start_color="FF00FF00", end_color="FF00FF00", patternType='solid')            
            feuille.cell(row = a[1], column = j).fill = fill

        wb.save(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')

    def retard_avion(self):
        #on récupère l'heure actuelle
        start_time = time.time()
        #on parcourt la colonne des heures de départ pour obtenir ceux qui ont déjà dû partir
        wb = xl.load_workbook('/Users/tom-b/OneDrive/Documents/Dossier centrale/cours 2A/PAi/PAI-3-Branche-Thibaut-1/vols.xlsxx')
        feuille = wb['Vols en cours']
        colonne_heure_départ = []
        for row in feuille.iter_rows():
            for cell in row :
                if not cell.value != None : 
                    a = (cell.coordinate[0],cell.coordinate[1])
                    print(a)
                    #liste de liste des coordonnées des cellules contenant une heure de départ
                    colonne_heure_départ.append(a[0])
                    colonne_heure_départ.append(a[1])
                    print(colonne_heure_départ)


        #récupération des heures d'arrivée normales
        #différence heure d'arrivée prévu/temps actuel
        #changement code couleur
             
    def message_arrive(self,corps,id_aeronef):              # Fonction terminée fonctionnelle
        
        # Identification de l'aeronef
        ligne=corps[0].split('-')
        b = ligne[1].split(' ')
        id = ' '+b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]+' '
        idbdd = b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]
        

        # Supression BDD

        conn = sqlite3.connect(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols_pai_3.db')
        cur = conn.cursor()

        cur.execute('''DELETE FROM "Plans de vols" WHERE Aeronef = ? ''', (idbdd,))

        conn.commit()
        conn.close()

        # Suppression ligne Excel

        wb = xl.load_workbook(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
        feuille = wb['Vols en cours']

        for row in feuille.iter_rows():
             for cell in row :
                 if str(cell.value) == str(id) : 
                    ligne = cell.row
                    
        for j in range(4,11):
            feuille.cell(ligne,j).value = None
            feuille.cell(ligne,j).fill = xl.styles.PatternFill(fill_type=None)


        wb.save(r'\Users\tom-b\OneDrive\Documents\Dossier centrale\cours 2A\PAi\PAi-3\PAI-3\vols.xlsx')
             
    def message_refus(self,corps,id_aeronef):
        #fonction à écrire
        a=True
        
    def message_acceptation(self,corps,id_aeronef):
        #fonction à écrire
        a=True

    def plan_de_vol_complementaire(self,corps,id_aeronef):
        #fonction à écrire
        a=True    

    def tri_geographique(self,corps,id_aeronef,decoupage) : 

        res = True

        # Connexion à la maessagerie

        ORG_EMAIL = "@outlook.fr" 
        usernm = "test.pai3" + ORG_EMAIL 
        passwd = "Tomblanchard3."
        conn = imaplib.IMAP4_SSL('outlook.office365.com')
        conn.login(usernm,passwd)
        conn.select('Inbox')

        # Recuperation arrivée, départ et noms de ville

        ligne=corps[4].split('-')
        depart=[ligne[1][1:5]]
        
        
        ligne2=corps[8].split('-')
        arrivee=[ligne2[1][1:5]]
        

        ligne3 = corps[6].split(' ')
        chemin = ligne3[2:len(ligne3)-1]
        

        liste_geo = depart + arrivee 
        
        # Est ce que les villes sont dans la zone de surveillance ? 

        #base = sqlite3.connect('/Users/thibautdejean/Downloads/PAI/Aerodromes.db')
        #cur = base.cursor()

        #for lieu in liste_geo : 
        #    cur.execute("SELECT ? FROM table  WHERE Aeronef = ? ", (decoupage,lieu))
        #   if cur.fechall() == 0:
        #        res = False
            

        # base.close()
        # Recherche de l'identifiant du mail
        if res == False : 
            typ, data = conn.search(None, '(OR SUBJECT ? BODY ?)',(id_aeronef,id_aeronef))

            for num in data[0].split():
            
                typ, msg_data = conn.fetch(num, '(RFC822)')
                msg = email.message_from_bytes(msg_data[0][1])

                if id_aeronef in msg.get_payload().lower():
                
                    msgnum = num.decode('utf-8')

        # Déplacement du mail       

        conn.copy(msgnum, 'Hors_zone')
        conn.store(msgnum, '+FLAGS', '\\Deleted')
        conn.expunge()

                
    ### Reconnaissance du type de mail ###
    def reconnaissance(self, corps):
        #exemple de mail :
        # corps='\n'
        # corps+='\n'
        # corps+='FF LFXOYWYX LFXVYWYX LFBWYWYX LFWBYWYX LFMMYWYX\n\n'
        # corps+='FF LFMYZPZX LFMUZPZX LFMUZTZX LFMIZPZX \n\n' 
        # corps+='310740 LFXOYXYX\n\n'
        # corps+='(FPL-COTE44-VM \n\n'
        # corps+='-SR20/L-GOLDURYB/S\n\n'
        # corps+='-LFMY1130\n\n'
        # corps+='-N0125A005 OAT UZES SAINT HIPPOLYTTE DU FORT BEDARIEUX\n\n' 
        # corps+='-LFMU 0120 LFMI LFMV\n\n'
        # corps+='-DOF/220831 REG/COYOTE 44 OPR/FAF RMK/NPL12MY)\n\n\n\n'
        
        #on récupère la partie du mail qui nous intéresse
        corps=corps.split("(")
        #séparation du mail ligne par ligne
        corps=corps[1].split('\n')
        #la première ligne nous permet de détecter le type de message
        ligne=corps[0].split('-')
        type_message=ligne[0].strip(' ')
        id_aeronef=ligne[1]
        decoupage = self.decoupage
        print(decoupage)
        #on regarde si le vole passe par une ville surveillée
        #  `self.tri_geographique(corps,id_aeronef,decoupage)

        #on envoie vers une fonction spécifique selon le type de mail :
        if type_message=='FPL':
            self.plan_de_vol(corps,id_aeronef)
            self.ecriture_excel(corps,id_aeronef)
            
        elif type_message=='DLA':
            self.message_delai(corps,id_aeronef)
        elif type_message=='CHG':
            self.message_changement(corps,id_aeronef)
        elif type_message=='CNL':
            self.message_annulation(corps,id_aeronef)
        elif type_message=='DEP':
            self.message_depart(corps,id_aeronef)
        elif type_message=='ARR':
            self.message_arrive(corps,id_aeronef)
        elif type_message=='REFUS':
            self.message_refus(corps,id_aeronef)
        elif type_message=='ACP':
            self.message_acceptation(corps,id_aeronef)
        elif type_message=='SPL':
            self.plan_de_vol_complementaire(corps,id_aeronef)


    def affichage_zone(self,event):
         self.boutonValider.config(state=ACTIVE)
         i = self.__zone.curselection()
         img = ",".join([self.__zone.get(j) for j in i])
         if img=="Plan NORM":
            self.__img = ImageTk.PhotoImage(Image.open('norm.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "NORM"
         elif img=="Plan LY00":
            self.__img = ImageTk.PhotoImage(Image.open('LY00.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "LY00"
         elif img=="Plan LY1T":
            self.__img = ImageTk.PhotoImage(Image.open('LY1T.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "LY1T"
         elif img=="Plan LY10":
            self.__img = ImageTk.PhotoImage(Image.open('LY10.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "LY10"
         elif img=="Plan LY11":
            self.__img = ImageTk.PhotoImage(Image.open('LY11.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "LY11"
         elif img=="Plan MM1L":
            self.__img = ImageTk.PhotoImage(Image.open('MM1L.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "MM1L"
         elif img=="Plan TR00":
            self.__img = ImageTk.PhotoImage(Image.open('TR00.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "TR00"
         elif img=="Plan TR10":
            self.__img = ImageTk.PhotoImage(Image.open('TR10.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "TR10"
         elif img=="Plan TR11":
            self.__img = ImageTk.PhotoImage(Image.open('TR11.png')) 
            self.__zoneAffichage.create_image(150, 145, image=self.__img)
            self.decoupage = "TR11"
            
    def __init__(self,base):
         global etat
         etat=True 
         Tk.__init__(self)
         # Configuration de la base de données
         self.__conn = sqlite3.connect(base)
         # on est connecté à la base de données
         curseur=self.__conn.cursor()
         ##Pour exécuter une requête :
         # curseur.execute("")
         # curseur.fetchall()[0][0]

         self.title('Surveillance trafic aérien')
         self.configure(bg="grey")
          
          # La barre d'outils composé de 2 boutons :
         self.__barreOutils = Frame(self)
         self.__barreOutils.pack(side=BOTTOM, padx=5, pady=5)
         self.boutonValider=Button(self.__barreOutils,text='Valider')
         self.boutonValider.pack(side=LEFT, padx=5, pady=5)
         self.boutonValider.config(state=DISABLED)
         self.__QuitButton = Button(self.__barreOutils, text ='Quitter', width=13)
         self.__QuitButton.pack(side=RIGHT, padx=5, pady=5)
         self.__ZC=Frame(self,borderwidth=2,relief=GROOVE,bg='white')
         self.__ZC.pack(side=TOP,padx=5,pady=2)
        # Configuration du Label de l'en-tête qui sert à donner des indications
         self.label_enTete=Label(self.__ZC,text="Définissez le découpage du territoire :", bg='white',fg="black",font=("Arial",15))
         self.label_enTete.pack(side=LEFT, padx=20,pady=8) 
        # Configuration du choix du découpage du territoire
         self.__zone = Listbox(self.__ZC)
         self.__zone.insert(1, "Plan NORM")
         self.__zone.insert(2, "Plan LY00")
         self.__zone.insert(3, "Plan LY1T")
         self.__zone.insert(4, "Plan LY10")
         self.__zone.insert(5, "Plan LY11")
         self.__zone.insert(6, "Plan MM1L")
         self.__zone.insert(7, "Plan TR00")
         self.__zone.insert(8, "Plan TR10")
         self.__zone.insert(9, "Plan TR11")
         self.__zone.pack(padx=20,pady=8) 
         self.__zone.bind('<<ListboxSelect>>',self.affichage_zone)



        # deuxième fenêtre :
         self._fenetre=Toplevel(self)
         self._fenetre.title("Vol en cours")
         self._fenetre.geometry("1355x700")
         self._fenetre.geometry("+0+0")
         self.__barreOutils2 = Frame(self._fenetre)
         self.__barreOutils2.pack(side=BOTTOM, padx=5, pady=5)
         self.__QuitButton2 = Button(self.__barreOutils2, text ='Quitter', width=13)
         self.__QuitButton2.pack(side=RIGHT, padx=5, pady=5)
         # Configuration du Label de l'en-tête qui sert à donner des indications
         self.__ZC2=Frame(self._fenetre,borderwidth=2,relief=GROOVE,bg='white')
         self.__ZC2.pack(side=TOP,padx=5,pady=2)
         self.label_enTete2=Label(self.__ZC2,text="Trafic aérien actuel : ", bg='white',fg="black",font=("Arial",15))
         self.label_enTete2.pack(side=LEFT, padx=20,pady=8) 
         self.__QuitButton2.config(command=self.fin2)
        # Le canvas pour afficher le plan du découpage
         self.__zoneAffichage =Canvas(self, width = 300,height = 290,bg='white')  
         self.__zoneAffichage.pack()
        # Commandes associées aux boutons
         self.__QuitButton.config(command=self.fin)
         #phase de test : 
         self.boutonValider.config(command=self.nouvelle_fenetre)
         
         
    def fin(self):
         global etat
         etat=False 
         self.destroy()
         raise SystemExit
         
    def fin2(self):
        global etat
        etat=False 
        self._fenetre.destroy()
        self.destroy()
        raise SystemExit
        
### Interface Graphique (affichage des vols en cours) ###

    def nouvelle_fenetre(self):
         global etat
         #Fermeture de la première fenêtre
         self.withdraw()
         # self._fenetre.deiconify()
         self.lecture_mail()
         
         
    ### définnir un cycle qui relance la fonction à un intervalle de temps précis afin de lire les nouveaux mails.
    def lecture_mail(self):
        global etat
        servername='outlook.office365.com'
        self._fenetre.deiconify()
        ### Lecture des mails ###
        start_time = time.time()
        interval = 10 #on récupère des nouveaux mails toutes les 10 secondes
        j=0
        while etat==True:
            j+=interval
            time.sleep(start_time + j - time.time())
            (i,data,conn)=connexion(servername)
            self.mail(i,data,conn)
    
    def mail(self,i,data,conn):
        #On parcours les mails 1 par 1.
        for x in range(i):
            latest_email_uid = data[0].split()[x]
            result, email_data = conn.uid('fetch', latest_email_uid, '(RFC822)')
            # result, email_data = conn.store(num,'-FLAGS','\\Seen') 
            # this might work to set flag to seen, if it doesn't already
            raw_email = email_data[0][1]
            raw_email_string = raw_email.decode('utf-8')
            email_message = email.message_from_string(raw_email_string)
        
            ### informations sur le sujet du mail :
            # Header Details
            # date_tuple = email.utils.parsedate_tz(email_message['Date'])
            # if date_tuple:
            #     local_date = datetime.datetime.fromtimestamp(email.utils.mktime_tz(date_tuple))
            #     local_message_date = "%s" %(str(local_date.strftime("%a, %d %b %Y %H:%M:%S")))
            # email_from = str(email.header.make_header(email.header.decode_header(email_message['From'])))
            # email_to = str(email.header.make_header(email.header.decode_header(email_message['To'])))
            # subject = str(email.header.make_header(email.header.decode_header(email_message['Subject'])))
        
            # Body details
            for part in email_message.walk():
                if part.get_content_type() == "text/plain":
                    body = part.get_payload(decode=True)
                    corps=body.decode('UTF-8')
                    print(corps)
                    self.reconnaissance(corps)
                    # on obtient de chaîne de caractère qu'il faut maintenan traiter.
                    
                    ###Possibilité de stocker chaque mail dans un fichier txt : 
                    # file_name = "email_" + str(x) + ".txt"
                    # output_file = open(file_name, 'w')
                    # output_file.write("From: %s\nTo: %s\nDate: %s\nSubject: %s\n\nBody: \n\n%s" %(email_from, email_to,local_message_date, subject, body.decode('utf-8')))
                    # output_file.close()
                    
                else:
                    continue
        print("mails traités, pas de mails non lus")
                
                
### Programme Principal ###
if __name__ == '__main__':
### Création de l'interface Graphique ###
    fen = FenPrincipale("nom_base.db")
    fen.mainloop()  